import random
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import torch
from torch.utils.data import Dataset, DataLoader
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torch.optim.lr_scheduler import ReduceLROnPlateau
from sklearn.metrics import confusion_matrix, classification_report
import math
from collections import Counter
import matplotlib.cm as cm
from matplotlib.collections import LineCollection
from matplotlib.lines import Line2D
import openpyxl
import json
import os
from matplotlib import gridspec
from mpl_toolkits.axes_grid1 import make_axes_locatable

# Reproducability
def set_seed(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

'''
RELEVANT TO DATASET CREATION PART
'''
# Create anomaly ranges function
def create_anomaly_ranges(y_labels):
    anomaly_ranges = {}
    current_label = y_labels[0]
    start_index = 0

    for i in range(1, len(y_labels)):
        if y_labels[i] != current_label:
            if current_label != 0:  # Only consider non-zero (anomaly) labels
                if current_label not in anomaly_ranges:
                    anomaly_ranges[current_label] = []
                anomaly_ranges[current_label].append((start_index, i - 1))
            current_label = y_labels[i]
            start_index = i

    # Add the last range if it is an anomaly
    if current_label != 0:
        if current_label not in anomaly_ranges:
            anomaly_ranges[current_label] = []
        anomaly_ranges[current_label].append((start_index, len(y_labels) - 1))

    return anomaly_ranges

# Define a function to create sliding windows with trace boundaries
def create_sliding_windows(traces, window_size=40, stride=1):
    windowed_traces = []

    for trace_index,trace in enumerate(traces):
        
        num_windows = (len(trace) - window_size) // stride + 1

        # Generate windows
        windows = np.array([trace[i:i + window_size] for i in range(0, num_windows * stride, stride)])
        windowed_traces.append(windows)

    return windowed_traces

def generate_labels_for_window(window_start, window_end, anomaly_ranges):
    label_window = []
    for i in range(window_start, window_end + 1):
        label = 0  # Default to 'normal'
        for anomaly_type, ranges in anomaly_ranges.items():
            for (range_start, range_end) in ranges:
                # Check if the point is within the anomaly range
                if range_start <= i <= range_end:
                    label = anomaly_type  # Within anomaly range
                    break  # Stop checking if label is set
            if label != 0:  # Exit if labeled with anomaly
                break
        label_window.append(label)

    return label_window

# Function to create sliding windows and assign labels for each trace
def generate_labels_for_traces(traces, y_traces, window_size, stride):
    traces_labels = []

    for trace, y_labels in zip(traces, y_traces):
        windows_labels = []
        anomaly_ranges = create_anomaly_ranges(y_labels)
        num_windows = (len(trace) - window_size) // stride + 1
        
        for i in range(num_windows):
            window_start = i * stride
            window_end = window_start + window_size - 1
            window_labels = generate_labels_for_window(window_start, window_end, anomaly_ranges)
            windows_labels.append(list(window_labels))
        traces_labels.append(windows_labels)
    return traces_labels


'''
RELEVANT TO CLASS WEIGHTS PART
'''
def count_labels(traces_windows_labels, num_classes):

    # Dictionary to hold the count for each label
    label_counts = {label: 0 for label in range(num_classes)}
    
    for trace_windows_labels in traces_windows_labels:
        for window_label in trace_windows_labels:
            for label in window_label:
                # Increment the count for the label
                if label in label_counts:
                    label_counts[label] += 1
                else:
                    label_counts[label] = 1
    
    return label_counts

def calculate_class_ratios(train_disturbed_windows_labels, num_classes):
    # Replace class_counts with counts from training data
    train_label_counts = count_labels(train_disturbed_windows_labels, num_classes)  # Count labels in the training dataset

    # Total number of samples
    total_samples = sum(train_label_counts.values())

    # Calculate weights for each class dynamically
    class_ratios = {cls: total_samples / count for cls, count in train_label_counts.items()}

    return class_ratios


def calculate_scaled_class_weights(train_disturbed_windows_labels, num_classes):
        # Replace class_counts with counts from training data
        train_label_counts = count_labels(train_disturbed_windows_labels, num_classes)  # Count labels in the training dataset

        # Total number of samples
        total_samples = sum(train_label_counts.values())

        # Calculate weights for each class dynamically
        class_weights = {cls: total_samples / count for cls, count in train_label_counts.items()}

        # Normalize weights to keep them in a reasonable range (optional)
        max_weight = 1#max(class_weights.values())
        class_weights = {cls: weight / max_weight for cls, weight in class_weights.items()}

        # Convert to a tensor
        weights_tensor = torch.tensor([class_weights[cls] for cls in train_label_counts.keys()])

        min_val = 0.3
        max_val = 1.0
        scaled_weights = (weights_tensor - weights_tensor.min()) / (weights_tensor.max() - weights_tensor.min())
        scaled_weights = scaled_weights * (max_val - min_val) + min_val

        return scaled_weights


'''
RELEVANT TO EVALUATION METRICS PART
'''
# Post-processing function to extract anomaly types and intervals from labels
def extract_anomalies_from_labels(previous_window_labels, window_labels, next_window_labels, num_classes):
    anomalies = {"anomaly_type": [], "anomaly_intervals": []}
    current_anomaly_type = None
    current_start = None
    current_end = None

    idx = 0
    while idx < len(window_labels):
        label = window_labels[idx]

        if idx == 0 and label != 0:  # if the first index labeled with anomaly 
            anomaly_code = window_labels[idx]  # Get the first character as anomaly type
            for key, value in range(num_classes):
                if value == anomaly_code:
                    current_anomaly_type = key
                    break
            current_start = idx  # Set start position

        if current_anomaly_type != None and current_start != None and current_end != None:
            # Append the anomaly type and interval
            anomalies["anomaly_type"].append(current_anomaly_type)
            anomalies["anomaly_intervals"].append((current_start, current_end))

            # Reset for the next anomaly
            current_anomaly_type = None
            current_start = None
            current_end = None

        idx += 1

    return anomalies

# Create anomaly ranges function
def extract_anomalies_from_trace(trace_labels):
    anomalies = {"anomaly_type": [], "anomaly_intervals": []}
    current_anomaly_type = None
    start_index = None

    for idx, label in enumerate(trace_labels):
        if label == 0:
            # If we were tracking an anomaly, finalize it
            if current_anomaly_type is not None:
                anomalies["anomaly_type"].append(current_anomaly_type)
                anomalies["anomaly_intervals"].append((start_index, idx - 1))
                current_anomaly_type = None
                start_index = None
        
        else:
            # Regular anomaly label
            if current_anomaly_type is None:
                current_anomaly_type = label
                start_index = idx
            elif current_anomaly_type != label:
                # Store previous anomaly and start a new one
                anomalies["anomaly_type"].append(current_anomaly_type)
                anomalies["anomaly_intervals"].append((start_index, idx - 1))
                current_anomaly_type = label
                start_index = idx

    # Handle last anomaly segment if the trace ends with an anomaly
    if current_anomaly_type is not None:
        anomalies["anomaly_type"].append(current_anomaly_type)
        anomalies["anomaly_intervals"].append((start_index, len(trace_labels) - 1))

    return anomalies


def extract_true_labels(traces_labels, stride, num_future_samples):
    all_true_labels = []
    for trace_labels in  traces_labels:
        trace_true_labels = []
        num_windows = len(trace_labels)

        for window_index in range(num_windows - num_future_samples):
         
            # Create true labels for current window
            trace_future_labels = trace_labels[window_index + stride : window_index + stride + num_future_samples]
            true_labels = []
           
            for window_future_labels in trace_future_labels:
                true_labels.extend(window_future_labels[-stride:]) 

            # Add results for current window to a trace
            if window_index == num_windows - num_future_samples - 1:
                trace_true_labels.extend(true_labels)
            else:
                trace_true_labels.extend(true_labels[:stride])

        # Store true and predicted labels for the entire trace
        all_true_labels.append(trace_true_labels)

    return all_true_labels

def save_undisturbed_results(path, model_name, true_labels, mode1_predicted_labels, mode2_predicted_labels, mode3_predicted_labels, mode4_predicted_labels):

    # Flatten lists
    true_flat = [item for sublist in true_labels for item in sublist]
    mode1_pred_flat = [item for sublist in mode1_predicted_labels for item in sublist]
    mode2_pred_flat = [item for sublist in mode2_predicted_labels for item in sublist]
    mode3_pred_flat = [item for sublist in mode3_predicted_labels for item in sublist] 
    mode4_pred_flat = [item for sublist in mode4_predicted_labels for item in sublist]

    # Get unique classes (only those that appear)
    classes = sorted(list(set(true_flat) | set(mode1_pred_flat) | set(mode2_pred_flat) | set(mode3_pred_flat) | set(mode4_pred_flat)))
    label_names = ['0', '1', '2', '3', '4', '5', '6', '7']

    # Confusion matrices
    cm_mode1_pred = confusion_matrix(true_flat, mode1_pred_flat, labels=classes)
    cm_mode2_pred = confusion_matrix(true_flat, mode2_pred_flat, labels=classes)
    cm_mode3_pred = confusion_matrix(true_flat, mode3_pred_flat, labels=classes)
    cm_mode4_pred = confusion_matrix(true_flat, mode4_pred_flat, labels=classes)

    # Only show first row (true label '0')
    normal_index = classes.index(0) if 0 in classes else None
    if normal_index is None:
        print("Label index for normal class not found in confusion matrix.")
        return

    row_mode1_pred = cm_mode1_pred[normal_index, :].reshape(1, -1)
    row_mode2_pred = cm_mode2_pred[normal_index, :].reshape(1, -1)
    row_mode3_pred = cm_mode3_pred[normal_index, :].reshape(1, -1)
    row_mode4_pred = cm_mode4_pred[normal_index, :].reshape(1, -1)

    # Plot the cropped confusion matrices
    fig, axes = plt.subplots(2, 2, figsize=(16, 6))

    sns.heatmap(row_mode1_pred, annot=True, fmt="d", cmap="Blues",
                xticklabels=label_names, yticklabels=['a'], ax=axes[0, 0])
    axes[0, 0].set_xlabel("Mode1 Predicted")
    axes[0, 0].set_ylabel("True")
    axes[0, 0].set_title("Mode1")

    sns.heatmap(row_mode2_pred, annot=True, fmt="d", cmap="Greens",
                xticklabels=label_names, yticklabels=['a'], ax=axes[0, 1])
    axes[0, 1].set_xlabel("Mode2 Predicted")
    axes[0, 1].set_ylabel("True")
    axes[0, 1].set_title("Mode2")

    sns.heatmap(row_mode3_pred, annot=True, fmt="d", cmap="Oranges",
                xticklabels=label_names, yticklabels=['a'], ax=axes[1, 0])
    axes[1, 0].set_xlabel("Mode3 Predicted")
    axes[1, 0].set_ylabel("True")
    axes[1, 0].set_title("Mode3")

    sns.heatmap(row_mode4_pred, annot=True, fmt="d", cmap="Purples",
                xticklabels=label_names, yticklabels=['a'], ax=axes[1, 1])
    axes[1, 1].set_xlabel("Mode4 Predicted")
    axes[1, 1].set_ylabel("True")
    axes[1, 1].set_title("Mode4")


    plt.tight_layout()

    # Save figure
    os.makedirs(path, exist_ok=True)
    file_path = os.path.join(path, f"{model_name}_undisturbed_confusion_matrix.png")
    plt.savefig(file_path)
    plt.close()


def save_binary_AD_metrics(model_name, filename, ad_levels, global_f_scores, global_precisions, global_recalls, types_wise_metrics):

    try:
        book = openpyxl.load_workbook(filename)
        file_exists = True
    except FileNotFoundError:
        book = None
        file_exists = False

    # Use 'a' mode if file exists, otherwise 'w' mode
    mode = 'a' if file_exists else 'w'
    if_sheet_exists_option = "replace" if file_exists else None

    with pd.ExcelWriter(filename, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists_option) as writer:
        for i, ad_level in enumerate(ad_levels):
            sheet_name = f"AD Level {ad_level}"

            # Prepare new data as a DataFrame
            new_data = {
                "Model Name": [model_name],
                "F1": [round(global_f_scores[i], 3)],
                "Precision": [round(global_precisions[i], 3)],
                "Recall": [round(global_recalls[i], 3)],
            }

            # Adding type-wise metrics in the requested format
            for anomaly_type in types_wise_metrics.keys():
                new_data[f"T{list(types_wise_metrics.keys()).index(anomaly_type) + 1} F1"] = [round(types_wise_metrics[anomaly_type][i]["f_score"], 3)]
                new_data[f"T{list(types_wise_metrics.keys()).index(anomaly_type) + 1} Recall"] = [round(types_wise_metrics[anomaly_type][i]["recall"], 3)]

            df_new = pd.DataFrame(new_data)

            # If the sheet exists, read existing data, append, and overwrite
            if file_exists and sheet_name in book.sheetnames:
                existing_df = pd.read_excel(filename, sheet_name=sheet_name)
                df_combined = pd.concat([existing_df, df_new], ignore_index=True)
            else:
                df_combined = df_new

            # Write the updated data to the Excel file
            df_combined.to_excel(writer, sheet_name=sheet_name, index=False)
        
        ad_weights = [0.4, 0.3, 0.2, 0.1]
        weighted_f1 = sum(w * f1 for w, f1 in zip(ad_weights, global_f_scores))
        weighted_precision = sum(w * p for w, p in zip(ad_weights, global_precisions))
        weighted_recall = sum(w * r for w, r in zip(ad_weights, global_recalls))

        summary_df = pd.DataFrame({
        "Model Name": [model_name],
        "F1": [round(weighted_f1, 3)],
        "Precision": [round(weighted_precision, 3)],
        "Recall": [round(weighted_recall, 3)]
        })

        # Add weighted type-wise metrics
        for anomaly_type in types_wise_metrics.keys():
            idx = list(types_wise_metrics.keys()).index(anomaly_type) + 1
            weighted_type_f1 = sum(w * types_wise_metrics[anomaly_type][i]["f_score"] for i, w in enumerate(ad_weights))
            weighted_type_recall = sum(w * types_wise_metrics[anomaly_type][i]["recall"] for i, w in enumerate(ad_weights))
            summary_df[f"T{idx} F1"] = [round(weighted_type_f1, 3)]
            summary_df[f"T{idx} Recall"] = [round(weighted_type_recall, 3)]

        # If the sheet exists, read existing data, append, and overwrite
        if file_exists and sheet_name in book.sheetnames:
            existing_df = pd.read_excel(filename, sheet_name="Overall AD Performance")
            df_combined = pd.concat([existing_df, summary_df], ignore_index=True)
        else:
            df_combined = summary_df

        df_combined.to_excel(writer, sheet_name="Overall AD Performance", index=False)


def save_multiclass_AD_metrics(model_name, filename, ad_levels, global_f_scores, global_precisions, global_recalls, types_wise_metrics):

    # Check if the file exists
    try:
        book = openpyxl.load_workbook(filename)
        file_exists = True
    except FileNotFoundError:
        book = None
        file_exists = False

    # Use 'a' mode if file exists, otherwise 'w' mode
    mode = 'a' if file_exists else 'w'
    if_sheet_exists_option = "replace" if file_exists else None

    with pd.ExcelWriter(filename, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists_option) as writer:
        for i, ad_level in enumerate(ad_levels):
            sheet_name = f"AD Level {ad_level}"

            # Prepare new data as a DataFrame
            new_data = {
                "Model Name": [model_name],
                "F1": [round(global_f_scores[i], 3)],
                "Precision": [round(global_precisions[i], 3)],
                "Recall": [round(global_recalls[i], 3)],
            }

            # Adding type-wise metrics in the requested format
            for anomaly_type in types_wise_metrics.keys():
                new_data[f"T{list(types_wise_metrics.keys()).index(anomaly_type) + 1} F1"] = [round(types_wise_metrics[anomaly_type][i]["f_score"], 3)]
                new_data[f"T{list(types_wise_metrics.keys()).index(anomaly_type) + 1} Precision"] = [round(types_wise_metrics[anomaly_type][i]["precision"], 3)]
                new_data[f"T{list(types_wise_metrics.keys()).index(anomaly_type) + 1} Recall"] = [round(types_wise_metrics[anomaly_type][i]["recall"], 3)]

            df_new = pd.DataFrame(new_data)

            # If the sheet exists, read existing data, append, and overwrite
            if file_exists and sheet_name in book.sheetnames:
                existing_df = pd.read_excel(filename, sheet_name=sheet_name)
                df_combined = pd.concat([existing_df, df_new], ignore_index=True)
            else:
                df_combined = df_new

            # Write the updated data to the Excel file
            df_combined.to_excel(writer, sheet_name=sheet_name, index=False)

        ad_weights = [0.4, 0.3, 0.2, 0.1]
        weighted_f1 = sum(w * f1 for w, f1 in zip(ad_weights, global_f_scores))
        weighted_precision = sum(w * p for w, p in zip(ad_weights, global_precisions))
        weighted_recall = sum(w * r for w, r in zip(ad_weights, global_recalls))

        summary_df = pd.DataFrame({
        "Model Name": [model_name],
        "F1": [round(weighted_f1, 3)],
        "Precision": [round(weighted_precision, 3)],
        "Recall": [round(weighted_recall, 3)]
        })

        # Add weighted type-wise metrics
        for anomaly_type in types_wise_metrics.keys():
            idx = list(types_wise_metrics.keys()).index(anomaly_type) + 1
            weighted_type_f1 = sum(w * types_wise_metrics[anomaly_type][i]["f_score"] for i, w in enumerate(ad_weights))
            weighted_type_recall = sum(w * types_wise_metrics[anomaly_type][i]["recall"] for i, w in enumerate(ad_weights))
            weighted_type_precision = sum(w * types_wise_metrics[anomaly_type][i]["precision"] for i, w in enumerate(ad_weights))
            summary_df[f"T{idx} F1"] = [round(weighted_type_f1, 3)]
            summary_df[f"T{idx} Recall"] = [round(weighted_type_recall, 3)]
            summary_df[f"T{idx} Precision"] = [round(weighted_type_precision, 3)]

        # If the sheet exists, read existing data, append, and overwrite
        if file_exists and sheet_name in book.sheetnames:
            existing_df = pd.read_excel(filename, sheet_name="Overall AD Performance")
            df_combined = pd.concat([existing_df, summary_df], ignore_index=True)
        else:
            df_combined = summary_df

        df_combined.to_excel(writer, sheet_name="Overall AD Performance", index=False)


'''
RELEVANT TO VISUALIZATION PART
'''
def plot_predictions_four_modes(
    true_labels,
    mode1_predicted_labels,
    mode2_predicted_labels,
    mode3_predicted_labels,
    mode4_predicted_labels,
    mode1_confidences,
    mode2_confidences,
    mode3_confidences,
    mode4_confidences
):
    num_traces = len(true_labels)
    fig, axes = plt.subplots(num_traces, 4, figsize=(24, 4 * num_traces))
    plt.rcParams.update({'font.size': 10})    

    if num_traces == 1:
        axes = [axes]

    cmap = cm.get_cmap('Reds')

    for i, (ax_row, true, pred, masked_pred, pred3, pred4, conf, masked_conf, conf3, conf4) in enumerate(
        zip(
            axes,
            true_labels,
            mode1_predicted_labels,
            mode2_predicted_labels,
            mode3_predicted_labels,
            mode4_predicted_labels,
            mode1_confidences,
            mode2_confidences,
            mode3_confidences,
            mode4_confidences
        )
    ):
        def plot_mode(ax, pred_labels, conf_values, mode_name):
            ax.plot(true, marker='o', linestyle='-', markersize=3, label="True Labels", color="green", alpha=0.7)
            pred_points = np.array([range(len(pred_labels)), pred_labels]).T.reshape(-1, 1, 2)
            pred_segments = np.concatenate([pred_points[:-1], pred_points[1:]], axis=1)
            lc_pred = LineCollection(pred_segments, cmap=cmap, norm=plt.Normalize(0, 1))
            lc_pred.set_array(np.array(conf_values))
            lc_pred.set_linewidth(3)
            ax.add_collection(lc_pred)
            ax.autoscale()
            ax.set_title(f"Trace {i} - {mode_name} Labels with Confidence")
            ax.set_xlabel("Index")
            ax.set_ylabel("Value")
            ax.grid(True)
            cbar = plt.colorbar(lc_pred, ax=ax)
            cbar.set_label("Confidence")
            legend_elements = [
                Line2D([0], [0], color='green', lw=2, label='True Labels'),
                Line2D([0], [0], color='red', lw=2, label=f'{mode_name} Labels'),
            ]
            ax.legend(handles=legend_elements, loc='upper right')

        plot_mode(ax_row[0], pred, conf, "Mode 1 Predicted")
        plot_mode(ax_row[1], masked_pred, masked_conf, "Mode 2 Predicted")
        plot_mode(ax_row[2], pred3, conf3, "Mode 3 Predicted")
        plot_mode(ax_row[3], pred4, conf4, "Mode 4 Predicted")

    plt.tight_layout()
    plt.show()


def visualize_prediction_probs(
    selected_modes,  # list of indices: e.g. [0, 2]
    mode1_probs, mode2_probs, mode3_probs, mode4_probs,
    anomaly_ranges, window_size,
    trace_index, start_index, end_index
):
    mode_names = ['Mode 1', 'Mode 2', 'Mode 3', 'Mode 4']
    mode_probs = [mode1_probs, mode2_probs, mode3_probs, mode4_probs]

    # Time steps and ticks
    start_range = start_index + window_size
    final_range = end_index + window_size
    time_steps = list(range(start_range, final_range + 1))
    xticks = np.arange(len(time_steps)) + 0.5
    xtick_labels = time_steps

    # Generate true labels
    true_labels = []
    for t in time_steps:
        label = 0
        if trace_index in anomaly_ranges:
            for anomaly_type in anomaly_ranges[trace_index]:
                for start, end in anomaly_ranges[trace_index][anomaly_type]:
                    if start <= t <= end:
                        label = int(anomaly_type)
                        break
                if label != 0:
                    break
        true_labels.append(label)

    # Determine number of selected modes
    num_modes = len(selected_modes)
    num_rows = int(np.ceil(num_modes / 2))

    # === MODE PROBABILITIES PLOT ===
    fig = plt.figure(figsize=(12, 3 * num_rows))  # IEEE one-column width
    gs = gridspec.GridSpec(
        num_rows, 5,
        width_ratios=[1, 0.03, 1, 0.03, 0.1],
        height_ratios=[1] * num_rows,
        hspace=0.35,
        wspace=0.35
    )

    cmap = 'rocket_r'

    for plot_index, mode_idx in enumerate(selected_modes):
        row = plot_index // 2
        col = (plot_index % 2) * 2  # Column 0 or 2 for plot
        ax = fig.add_subplot(gs[row, col])
        cax = None

        probs_list = [np.array(mode_probs[mode_idx][trace_index][j]) for j in range(start_index, end_index + 1)]
        heatmap_data = np.array(probs_list).T  # (classes, time)

        num_classes = heatmap_data.shape[0]
        tick_positions = np.arange(num_classes) + 0.5
        tick_labels = reversed(['Normal','T1','T2','T3','T4','T5','T6','T7'])

        sns.heatmap(
            heatmap_data[::-1],
            annot=False,
            fmt=".2f",
            cmap=cmap,
            vmin=0.0,
            vmax=1.0,
            linewidths=0.4,
            linecolor='white',
            cbar=False,
            ax=ax
        )

        #ax.set_xticks(xticks)
        #ax.set_xticklabels(xtick_labels, rotation=90, fontsize=8)
        ax.set_xticks([])
        ax.set_xticklabels([])
        ax.set_xlabel("Time Index", fontsize=14)
        ax.set_yticks(tick_positions)
        ax.set_yticklabels(tick_labels, rotation=0, fontsize=12)
        ax.set_xlabel("Time Index", fontsize=14)
        ax.set_ylabel("Class", fontsize=14)
        ax.set_title(f'{mode_names[mode_idx]}', fontsize=16)

        for x in range(len(time_steps)):
            ax.axvline(x, color='gray', linewidth=0.2, alpha=0.3)

    # Add axis manually at far right
    cbar_ax = fig.add_axes([0.82, 0.15, 0.015, 0.7])  # [left, bottom, width, height]
    norm = plt.Normalize(0, 1)
    sm = plt.cm.ScalarMappable(cmap='rocket_r', norm=norm)
    sm.set_array([])
    fig.colorbar(sm, cax=cbar_ax, orientation='vertical', label="Probability")
    cbar_ax.set_ylabel('Probability', fontsize=14)

    # Save plot
    # Format selected mode indices with zero-padding
    mode_str = "_".join([f"{i:01d}" for i in selected_modes])

    # Build filename
    filename = f"../results/trace{trace_index}_index{start_range}_{final_range}_mode{mode_str}_prediction_probs.pdf"
    plt.savefig(filename, dpi=300, bbox_inches='tight', format='pdf')

    # === TRUE LABELS PLOT (SEPARATE, NOT SAVED) ===
    fig_label, ax_label = plt.subplots(figsize=(4, 4))  # IEEE column width

    ax_label.plot(time_steps, true_labels, color='black', linewidth=2, marker='o')
    ax_label.set_ylim(-0.5, max(true_labels) + 0.5)
    ax_label.set_yticks(sorted(set(true_labels)))
    ax_label.set_yticklabels(sorted(set(true_labels)), fontsize=9)
    ax_label.set_ylabel("True Label", fontsize=10)
    ax_label.set_xlabel("Time Index", fontsize=10)
    ax_label.set_title("True Class Labels", fontsize=11)
    ax_label.grid(True, linestyle='--', alpha=0.5)
    ax_label.set_xticks(time_steps)
    ax_label.set_xticklabels(time_steps, rotation=90, fontsize=8)

    plt.tight_layout()
    plt.show()





    


